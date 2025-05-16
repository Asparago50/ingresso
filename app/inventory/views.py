# ingresso/EntrataMerci/app/inventory/views.py

from django.shortcuts import render, redirect, get_object_or_404 # Aggiunto get_object_or_404 per DeleteView/UpdateView
from django.http import HttpResponse
from .models import Articolo
from .forms import ArticoloForm, ImportForm, ExportForm, MapColumnsForm
from django.contrib import messages
from .resources import ArticoloResource
from tablib import Dataset
import openpyxl
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy

# === IMPORT PER AUTENTICAZIONE E PERMESSI ===
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
# ===========================================

# --- Class-Based Views ---
class ArticoloListView(LoginRequiredMixin, PermissionRequiredMixin, ListView): # << AGGIUNTO MIXINS
    model = Articolo
    template_name = 'inventory/list.html'
    context_object_name = 'articoli'
    permission_required = 'inventory.view_articolo' # Permesso per visualizzare

    def get_queryset(self):
        return Articolo.objects.all().order_by('numero_sequenziale')

class ArticoloCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView): # << AGGIUNTO MIXINS
    model = Articolo
    form_class = ArticoloForm
    template_name = 'inventory/form.html'
    success_url = reverse_lazy('inventory:articolo_list')
    permission_required = 'inventory.add_articolo' # Permesso per aggiungere

    def form_valid(self, form):
        messages.success(self.request, "Articolo creato con successo!")
        return super().form_valid(form)

class ArticoloUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView): # << AGGIUNTO MIXINS
    model = Articolo
    form_class = ArticoloForm
    template_name = 'inventory/form.html'
    success_url = reverse_lazy('inventory:articolo_list')
    permission_required = 'inventory.change_articolo' # Permesso per modificare

    def form_valid(self, form):
        messages.success(self.request, "Articolo aggiornato con successo!")
        return super().form_valid(form)

class ArticoloDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView): # << AGGIUNTO MIXINS
    model = Articolo
    template_name = 'inventory/confirm_delete.html' # Dovrai creare questo template
    success_url = reverse_lazy('inventory:articolo_list')
    permission_required = 'inventory.delete_articolo' # Permesso per eliminare

    def form_valid(self, form):
        messages.success(self.request, "Articolo eliminato con successo!")
        return super().form_valid(form)


# --- Function-Based Views (Import/Export) ---
@login_required # << ORA DOVREBBE FUNZIONARE
@permission_required('inventory.add_articolo', raise_exception=True) # O 'inventory.change_articolo' se l'import aggiorna
def import_articoli(request):
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            # Salva temporaneamente il file se necessario o leggilo direttamente in memoria
            # (openpyxl può leggere da un file-like object)
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                # Mappatura delle colonne (esempio semplice, potresti renderlo più flessibile)
                # Questa è una mappatura di default, se non viene fornita dal form MapColumnsForm
                column_map_default = {
                    'A': 'numero_sequenziale',
                    'B': 'bolla_ddt_ft',
                    'C': 'data_bolla_ddt_ft',
                    'D': 'descrizione',
                    'E': 'quantita',
                    'F': 'unita_misura',
                    'G': 'note',
                }
                # Qui dovresti integrare la logica di MapColumnsForm se l'hai implementata
                # Per ora, usiamo la mappatura di default

                # Pulisci i dati esistenti se questa è la logica desiderata (ATTENZIONE!)
                # Articolo.objects.all().delete() # Scommenta con cautela!

                dataset = Dataset()
                # Salta l'intestazione se presente
                header = [cell.value for cell in sheet[1]] # Leggi l'intestazione
                # Potresti voler mappare l'intestazione ai campi del modello qui
                
                # Inizia a leggere i dati dalla seconda riga
                data_to_create = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
                    row_data = {}
                    has_data = False
                    for col_idx, cell in enumerate(row):
                        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                        field_name = column_map_default.get(col_letter)
                        if field_name:
                            row_data[field_name] = cell.value
                            if cell.value is not None and str(cell.value).strip() != "":
                                has_data = True
                    
                    if has_data: # Solo se la riga contiene dati
                        # Converti la data se necessario
                        if 'data_bolla_ddt_ft' in row_data and isinstance(row_data['data_bolla_ddt_ft'], str):
                            try:
                                from datetime import datetime
                                row_data['data_bolla_ddt_ft'] = datetime.strptime(row_data['data_bolla_ddt_ft'], '%Y-%m-%d').date()
                            except ValueError:
                                # Gestisci l'errore o imposta a None
                                row_data['data_bolla_ddt_ft'] = None
                        
                        # Se numero_sequenziale è vuoto, gestiscilo (es. errore o autoincremento se il DB lo permette e il campo non è unique)
                        if not row_data.get('numero_sequenziale'):
                            messages.error(request, f"Numero sequenziale mancante alla riga {row_idx + 2}. Riga saltata.")
                            continue # Salta questa riga

                        # Crea l'oggetto Articolo, o aggiornalo se esiste già
                        # Qui la logica di update_or_create è più complessa se numero_sequenziale non è PK
                        # Assumendo che numero_sequenziale sia univoco e vuoi aggiornare se esiste
                        try:
                            obj, created = Articolo.objects.update_or_create(
                                numero_sequenziale=row_data.get('numero_sequenziale'),
                                defaults=row_data
                            )
                            if created:
                                messages.info(request, f"Creato articolo: {obj.numero_sequenziale}")
                            else:
                                messages.info(request, f"Aggiornato articolo: {obj.numero_sequenziale}")
                        except Exception as e:
                            messages.error(request, f"Errore importando articolo con sequenziale {row_data.get('numero_sequenziale')}: {e}")

                messages.success(request, "Importazione completata con successo (o parzialmente). Controlla i messaggi.")
                return redirect('inventory:articolo_list')
            except Exception as e:
                messages.error(request, f"Errore durante l'importazione del file: {e}")
        else:
            messages.error(request, "Form non valido.")
    else:
        form = ImportForm()
    return render(request, 'inventory/form_import.html', {'form': form, 'page_title': 'Importa Articoli da Excel'})


@login_required # << ORA DOVREBBE FUNZIONARE
@permission_required('inventory.view_articolo', raise_exception=True) # O un permesso specifico per l'export
def export_articoli(request):
    articolo_resource = ArticoloResource()
    dataset = articolo_resource.export()
    response = HttpResponse(dataset.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="articoli.xlsx"'
    return response

# Se hai una vista per map_columns, proteggila anche:
# @login_required
# @permission_required('inventory.add_articolo', raise_exception=True)
# def map_columns_view(request):
#     ...